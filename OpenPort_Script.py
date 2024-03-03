import socket
import openpyxl

def check_port_status(ip, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(1)

    try:
        sock.connect((ip, port))
        return "Open"
    except socket.timeout:
        return "Filtered"
    except socket.error:
        return "Closed"
    finally:
        sock.close()

def check_ports_for_excel(input_excel_path, output_excel_path):
    wb = openpyxl.load_workbook(input_excel_path)
    ws = wb.active

    # Add headings to output sheet
    output_ws = wb.create_sheet("SinglePort")
    output_ws.append(["IP", "Port", "Port Status"])

    # Create a dictionary to store results per IP
    ip_results = {}
    cnt=1
    for row in ws.iter_rows(min_row=2, values_only=True):
        ip, ports_str = row

        if isinstance(ports_str, int):
            # Single port for the IP
            port_status = check_port_status(ip, ports_str)
            output_ws.append([ip, ports_str, port_status])
        else:
            ports = [int(port) for port in ports_str.split(",")]

            # List to store status for each port
            port_statuses = []

            for port in ports:
                port_status = check_port_status(ip, port)
                port_statuses.append(f"{port}/{port_status}")

            # Append the port statuses to the existing results for the IP
            
            if ip in ip_results:
                ip1= f"{ip}:{cnt}"
                ip_results[ip1].extend(port_statuses)
                #continue
            else:
                ip1= f"{ip}:{cnt}"
                ip_results[ip1] = port_statuses
        cnt+=1

    # Create output worksheet and add headings
    output_ws = wb.create_sheet("MultiplePort")
    output_ws.append(["IP", "Port/Status"])
    
    #for (ip, statuses) in enumerate(ip_results.items()):
    #    output_ws.append([f"{ip}", ', '.join(statuses)])
    for ip, statuses in ip_results.items():
        output_ws.append([ip, ', '.join(statuses)])
    wb.save(output_excel_path)

# Example usage with input Excel file and output Excel file
input_excel_path = "OpenPort.xlsx"
output_excel_path = "output_combined.xlsx"

check_ports_for_excel(input_excel_path, output_excel_path)
